from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable, List

from openpyxl import load_workbook
from sqlalchemy import Column, Date, Integer, MetaData, Numeric, String, Table, create_engine


DEFAULT_EXCEL_PATH = Path("database/Cabeçalho_Social.xlsx")
DEFAULT_DB_PATH = Path("database/classificacao_social_excel.db")


def get_headers(first_row: Iterable[object]) -> List[str]:
    return [str(cell).strip() for cell in first_row if cell not in (None, "")]


def guess_sqlalchemy_type(column_name: str):
    normalized = column_name.lower()

    if normalized in {"id", "pk", "codigo"}:
        return Integer
    if normalized.startswith("fk_") or normalized.startswith("id_"):
        return Integer
    if "valor" in normalized:
        return Numeric(15, 2)
    if "data" in normalized:
        return Date
    if any(token in normalized for token in ["total", "meta", "resultado", "pessoas", "prazo"]):
        return Integer
    if any(token in normalized for token in ["cpf", "cnpj", "cep", "contato", "email", "rg"]):
        return String(30)
    return String(255)


def build_schema(excel_path: Path, metadata: MetaData) -> list[Table]:
    wb = load_workbook(excel_path, data_only=True)
    created_tables: list[Table] = []

    for ws in wb.worksheets:
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = get_headers(first_row)

        if not headers:
            continue

        columns = []
        for header in headers:
            column_type = guess_sqlalchemy_type(header)
            is_pk = header.strip().lower() == "id"
            columns.append(Column(header, column_type, primary_key=is_pk))

        table = Table(ws.title, metadata, *columns)
        created_tables.append(table)

    return created_tables


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Cria um banco SQLite a partir do cabeçalho de abas de um arquivo Excel usando SQLAlchemy."
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=DEFAULT_EXCEL_PATH,
        help=f"Caminho do Excel de origem (padrão: {DEFAULT_EXCEL_PATH})",
    )
    parser.add_argument(
        "--db",
        type=Path,
        default=DEFAULT_DB_PATH,
        help=f"Caminho do arquivo SQLite de saída (padrão: {DEFAULT_DB_PATH})",
    )
    parser.add_argument(
        "--replace",
        action="store_true",
        help="Remove o banco existente antes de criar novamente.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    excel_path = args.excel
    db_path = args.db

    if not excel_path.exists():
        raise FileNotFoundError(f"Arquivo Excel não encontrado: {excel_path}")

    db_path.parent.mkdir(parents=True, exist_ok=True)

    if args.replace and db_path.exists():
        db_path.unlink()

    engine = create_engine(f"sqlite:///{db_path.as_posix()}")
    metadata = MetaData()

    tables = build_schema(excel_path, metadata)
    metadata.create_all(engine)

    print(f"Banco criado com sucesso: {db_path}")
    print(f"Total de tabelas criadas: {len(tables)}")
    for table in tables:
        print(f"- {table.name} ({len(table.columns)} colunas)")


if __name__ == "__main__":
    main()
